import datetime
import openpyxl
import io
from celery import shared_task
from apps.access.models import StudentReportFile, StudentReportDetail, User
from apps.my_learnings.helpers import get_one_year_datetime_from_now
from apps.web_portal.models.notification import Notification
from rest_framework.decorators import api_view
from django.http import JsonResponse
import json

@shared_task
def generate_report(report_detail_id):
    from apps.my_learnings.models import StudentEnrolledCourseTracker
    report_detail = StudentReportDetail.objects.get(pk=report_detail_id)
    print("Report Generation - Working", flush=True)

    if report_detail.content_group:
        content_group = report_detail.content_group
        user_groups = content_group.user_group.all()
    else:
        user_groups = None
    if report_detail.user_group:
        if user_groups and report_detail.user_group in user_groups:
            user_group = report_detail.user_group
            students = user_group.students.all()
        else:
            students = report_detail.user_group.students.all()
    else:
        students = User.objects.filter(
            created_by=report_detail.created_by
        )
    students = students.filter(
        created__date__gte=report_detail.start_date,
        created__date__lte=report_detail.end_date
    )
    for parameter in report_detail.filter_parameters.all():
        if parameter.identity == "Student Enrollment":
            enrolled_student_ids = StudentEnrolledCourseTracker.objects.filter(created_by__in=students).values_list('created_by_id', flat=True)
            students = students.filter(id__in=enrolled_student_ids)
        elif parameter.identity == "Student Course Progress":
            progress_student_ids = StudentEnrolledCourseTracker.objects.exclude(progress=0).filter(created_by__in=students).values_list('created_by_id', flat=True)
            students = students.filter(id__in=progress_student_ids)

    # Create a workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Name'
    sheet['B1'] = 'Email'
    sheet['C1'] = 'Birth Date'
    sheet['D1'] = 'Gender'
    sheet['E1'] = 'Identification Type'
    sheet['F1'] = 'Identification Number'
    sheet['G1'] = 'Admission ID'
    sheet['H1'] = 'Address'
    sheet['I1'] = 'City'   
    sheet['J1'] = 'State'
    sheet['K1'] = 'Country'
    sheet['L1'] = 'Pincode'
    sheet['M1'] = 'User Group'
    sheet['N1'] = 'Content Group'
    sheet['O1'] = 'Course Name'
    sheet['P1'] = 'Course Progress(%)'
    sheet['Q1'] = 'Enrolled Date'
    sheet['R1'] = 'Started On'
    sheet['S1'] = 'Completed On'

    row = 2
    for student in students:
        # birth_date isinstance of datetime.date so we have to format the date
        formatted_birth_date = student.birth_date.isoformat() if isinstance(student.birth_date, datetime.date) else None
        student_enrolled_courses = StudentEnrolledCourseTracker.objects.filter(created_by=student)
        if student_enrolled_courses.exists():
            for course in student_enrolled_courses:
                sheet.cell(row=row, column=1, value=student.first_name)
                sheet.cell(row=row, column=2, value=student.idp_email if student.idp_email else "None")
                sheet.cell(row=row, column=3, value=formatted_birth_date)
                sheet.cell(row=row, column=4, value=student.gender.identity if student.gender else "None")
                sheet.cell(row=row, column=5, value=student.identification_type.identity if student.identification_type else "None")
                sheet.cell(row=row, column=6, value=student.identification_number if student.identification_number else "None")
                sheet.cell(row=row, column=7, value=student.admission_id if student.admission_id else "None")
                sheet.cell(row=row, column=8, value=student.address if student.address else "None")
                sheet.cell(row=row, column=9, value=student.city.identity if student.city else "None")
                sheet.cell(row=row, column=10, value=student.state.identity if student.state else "None")
                sheet.cell(row=row, column=11, value=student.country.identity if student.country else "None")
                sheet.cell(row=row, column=12, value=student.pincode if student.pincode else "None")
                sheet.cell(row=row, column=13, value=report_detail.user_group.identity if report_detail.user_group else "None")
                sheet.cell(row=row, column=14, value=report_detail.content_group.identity if report_detail.content_group else "None")
                sheet.cell(row=row, column=15, value=course.entity.identity if course.entity else "None")
                sheet.cell(row=row, column=16, value=f"{course.progress}%")
                sheet.cell(row=row, column=17, value=f"{course.created}")
                sheet.cell(row=row, column=18, value=f"{course.started_on}")
                sheet.cell(row=row, column=19, value=f"{course.completed_on}")
                row += 1
        else:
            sheet.cell(row=row, column=1, value=student.first_name)
            sheet.cell(row=row, column=2, value=student.idp_email if student.idp_email else "None")
            sheet.cell(row=row, column=3, value=formatted_birth_date)
            sheet.cell(row=row, column=4, value=student.gender.identity if student.gender else "None")
            sheet.cell(row=row, column=5, value=student.identification_type.identity if student.identification_type else "None")
            sheet.cell(row=row, column=6, value=student.identification_number if student.identification_number else "None")
            sheet.cell(row=row, column=7, value=student.admission_id if student.admission_id else "None")
            sheet.cell(row=row, column=8, value=student.address if student.address else "None")
            sheet.cell(row=row, column=9, value=student.city.identity if student.city else "None")
            sheet.cell(row=row, column=10, value=student.state.identity if student.state else "None")
            sheet.cell(row=row, column=11, value=student.country.identity if student.country else "None")
            sheet.cell(row=row, column=12, value=student.pincode if student.pincode else "None")
            sheet.cell(row=row, column=13, value=report_detail.user_group.identity if report_detail.user_group else "None")
            sheet.cell(row=row, column=14, value=report_detail.content_group.identity if report_detail.content_group else "None")
            row += 1

    # Create a BytesIO object to store the workbook
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0) # 0: sets the reference point at the beginning of the file

    report_file = StudentReportFile.objects.create(created_by=report_detail.created_by)
    report_file.file.save(f'{report_detail.identity}.xlsx', output) # upload file name and file content

    report_detail.report = report_file
    report_detail.save()

@shared_task
def handle_user_enrolled_to_subscription_plan(user, subscription_plan_courses):
    """Handle the fact that user has enrolled to certification path. Creates trackers."""
    from apps.my_learnings.models import UserSubscriptionPlanTracker

    UserSubscriptionPlanTracker.objects.filter(created_by=user, entity=subscription_plan_courses).delete()

    tracker = UserSubscriptionPlanTracker.objects.create(
        created_by=user, entity=subscription_plan_courses, valid_till=get_one_year_datetime_from_now()
    )
    # chain of responsibility
    tracker.handle_user_enrolled()

    # notification = Notification(user=user, course=subscription_plan_courses, purpose="completed", message=f'Successfully enrolled in the course {subscription_plan_courses.identity}')
    # notification.save()
@shared_task
def generate_html_content(webinar_title, start_date, end_date):
    from datetime import timezone
    # Format the dates and times in a way that is compatible with the iCalendar format
    start_datetime = datetime.datetime.strptime(start_date, '%Y-%m-%dT%H:%M:%S.%fZ')
    end_datetime = datetime.datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%S.%fZ')

    # Construct the iCalendar event
    icalendar_event = f'''BEGIN:VCALENDAR
    VERSION:2.0
    PRODID:-//Techademy//NONSGML Webinar Reminder//EN
    BEGIN:VEVENT
    UID:webinar-{start_datetime}-{end_datetime}@techademy.com
    DTSTAMP:{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}
    DTSTART:{start_datetime}
    DTEND:{end_datetime}
    SUMMARY:{webinar_title}
    DESCRIPTION: Reminder for the webinar: {webinar_title}
    END:VEVENT
    END:VCALENDAR'''

    # Include the iCalendar event in the HTML content
    html_content = f'''
    <h4>Dear User,</h4>
    <p>Just a quick heads-up about our upcoming webinar "{webinar_title}"!</p>
    <p>Here are the details:</p>
    <ul>
        <li>Start Date: {start_date}</li>
        <li>End Date: {end_date}</li>
    </ul>
    <p>Add this webinar to your calendar: <a href="data:text/calendar;charset=utf-8,{icalendar_event}" download="{webinar_title}.ics">Download Event</a></p>
    '''

    return html_content

@shared_task
@api_view(['POST'])
def webinar_reminder_email(request):
    print("Reminder email")
    from apps.webinars.models import WebinarRegistration
    from datetime import date
    from apps.common.helpers import send_webinar_reminder_email
    from datetime import timedelta

    # Get WebinarRegistrations for today
    registrations = WebinarRegistration.objects.filter(
        status="success", 
        webinar__start_date=date.today()
    )
    session_start_times = []
    session_end_times = []
    if len(registrations) > 0:
        for registration in registrations:
            # Fetch related Webinar
            webinar = registration.webinar
            email = registration.user.idp_email
            # Fetch session details
            session_details = webinar.session_detail
            for sessions in session_details.values():
                for session in sessions:
                    session_start_times.append(session['start_time'])
                    session_end_times.append(session['end_time'])

            # Take the earliest start time and latest end time
            start_time = min(session_start_times)
            end_time = max(session_end_times)
            start_time_utc_dt = datetime.datetime.strptime(start_time, "%Y-%m-%dT%H:%M:%S.%fZ")
            end_time_utc_dt = datetime.datetime.strptime(end_time, "%Y-%m-%dT%H:%M:%S.%fZ")

            # Add 5 hours and 30 minutes to convert to IST
            ist_offset = timedelta(hours=5, minutes=30)
            start_time_ist = start_time_utc_dt + ist_offset
            end_time_ist = end_time_utc_dt + ist_offset

            # Format IST datetime objects as strings
            start_time_ist_str = start_time_ist.strftime("%Y-%m-%dT%H:%M:%S")
            end_time_ist_str = end_time_ist.strftime("%Y-%m-%dT%H:%M:%S")
            # Construct email content
            subject = "Friendly Reminder: Join Our Webinar!"
            html_content = f'<h4>Dear User, Just a quick heads-up about our upcoming webinar "{webinar.identity}"! We will be sharing valuable insights that could make a real difference for you. Here are the details you will need to access the webinar:<br>Start Time: {start_time_ist_str} <br> End Time: {end_time_ist_str} <br> Webinar Link: {webinar.webinar_link}</h4>'
            # html_content = generate_html_content(webinar.identity, start_time, end_time)
            # Send reminder email
            success = send_webinar_reminder_email("techcyces@gmail.com", subject, html_content, webinar.identity, start_time, end_time)
            
            # Prepare response
            # response_data = {
            #     "success": "Email sent successfully" if success else "Failed to send email",
            # }
            
            # return JsonResponse(response_data)
        
        return JsonResponse({"success": False, "message": "Email sent successfully"})
    else:
        return JsonResponse({"success": False, "message": "No webinar found for today"})